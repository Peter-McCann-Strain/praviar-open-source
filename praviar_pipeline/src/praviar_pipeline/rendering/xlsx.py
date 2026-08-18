"""XLSX export — claim charts and risk matrix for attorney workflows.

Produces a multi-sheet workbook:
  Sheet 1: Workbook Summary (visible brand and report metadata)
  Sheet 2: Risk Matrix (all patents, sorted by risk)
  Sheet 3+: Per-patent claim charts (one sheet per patent with claim chart data)
  Final sheet: Prior Art References
"""

from __future__ import annotations

import hashlib
import io
import re
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from praviar_pipeline.config import get_settings
from praviar_pipeline.models.analysis import RiskLevel
from praviar_pipeline.models.report_common import REPORT_DISCLAIMER
from praviar_pipeline.rendering.audience_projection import (
    AUDIENCE_PROJECTION_SCHEMA_VERSION,
    AudienceField,
)
from praviar_pipeline.rendering.brand_mark import render_praviar_mark_png_stream
from praviar_pipeline.rendering.branding import (
    SUPPORTED_OFFICE_BRANDING_LOGO_EXTENSIONS,
    format_artifact_title,
    resolve_branding_logo_path,
)
from praviar_pipeline.rendering.design import (
    BRAND_COPPER,
    BRAND_INK,
    BRAND_MINT,
    BRAND_PAPER,
    BRAND_SECONDARY_TEXT,
    BRAND_TEAL,
    RISK_BG,
    RISK_FILL,
)
from praviar_pipeline.rendering.evidence_scope import (
    source_status_detail,
    source_status_label,
    summarize_evidence_scope,
)
from praviar_pipeline.rendering.governed_decision import (
    governed_blocking_count,
    governed_decision_label,
    governed_patent_basis,
    governed_patent_posture,
)
from praviar_pipeline.rendering.spreadsheet_safety import (
    neutralize_spreadsheet_structure,
    neutralize_workbook_strings,
)

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from praviar_pipeline.models.report import FTOReport
    from praviar_pipeline.rendering.branding import BrandingConfig
    from praviar_pipeline.rendering.export_options import ExportRenderOptions


def _xlsx_hex(hex_color: str) -> str:
    return hex_color.lstrip("#").upper()


# Style constants
_HEADER_FILL = PatternFill(
    start_color=_xlsx_hex(BRAND_INK),
    end_color=_xlsx_hex(BRAND_INK),
    fill_type="solid",
)
_HEADER_FONT_PAPER = Font(bold=True, size=11, color=_xlsx_hex(BRAND_PAPER))
_RISK_FILLS = {
    risk_level.value: PatternFill(
        start_color=_xlsx_hex(RISK_FILL[risk_level]),
        end_color=_xlsx_hex(RISK_FILL[risk_level]),
        fill_type="solid",
    )
    for risk_level in RiskLevel
}
_DISCLOSED_FILLS = {
    "yes": PatternFill(
        start_color=_xlsx_hex(RISK_BG[RiskLevel.HIGH]),
        end_color=_xlsx_hex(RISK_BG[RiskLevel.HIGH]),
        fill_type="solid",
    ),
    "partial": PatternFill(
        start_color=_xlsx_hex(RISK_BG[RiskLevel.MEDIUM]),
        end_color=_xlsx_hex(RISK_BG[RiskLevel.MEDIUM]),
        fill_type="solid",
    ),
    "no": PatternFill(
        start_color=_xlsx_hex(RISK_BG[RiskLevel.CLEAR]),
        end_color=_xlsx_hex(RISK_BG[RiskLevel.CLEAR]),
        fill_type="solid",
    ),
}
_WRAP = Alignment(wrap_text=True, vertical="top")
_INVALID_SHEET_TITLE_CHARS = re.compile(r"[:\\/?*\[\]]")


def _safe_sheet_title(prefix: str, identifier: str, used_titles: set[str]) -> str:
    """Return an Excel-safe, stable title with collision-resistant truncation."""
    raw_title = f"{prefix}_{identifier}".strip()
    sanitized = _INVALID_SHEET_TITLE_CHARS.sub("_", raw_title).replace("'", "_").strip()
    if not sanitized:
        sanitized = prefix

    digest = hashlib.sha256(raw_title.encode("utf-8")).hexdigest()[:6].upper()

    if len(sanitized) > 31:
        keep = 31 - len(digest) - 1
        sanitized = f"{sanitized[:keep].rstrip()}_{digest}"

    candidate = sanitized
    counter = 2
    while candidate in used_titles:
        suffix = f"_{counter}"
        keep = 31 - len(suffix)
        candidate = f"{sanitized[:keep].rstrip()}{suffix}"
        counter += 1

    used_titles.add(candidate)
    return candidate


def _resolve_branding(branding: BrandingConfig | None) -> BrandingConfig:
    """Return caller branding or the canonical Praviar default."""
    from praviar_pipeline.rendering.branding import get_default_branding

    return branding or get_default_branding()


def _workbook_title(branding: BrandingConfig) -> str:
    """Return the visible workbook title for the selected brand context."""
    return str(format_artifact_title(branding, "FTO Workbook"))


def _sheet_branding_title(
    report: FTOReport,
    sheet_label: str,
    branding: BrandingConfig | None = None,
) -> str:
    """Return compact, deterministic worksheet branding for headers."""
    branding = _resolve_branding(branding)
    return f"{_workbook_title(branding)} | {report.compound.name} | {sheet_label}"


def _apply_working_sheet_branding(
    ws: Worksheet,
    report: FTOReport,
    *,
    header_row: int = 1,
    sheet_label: str,
    tab_color: str = BRAND_TEAL,
    branding: BrandingConfig | None = None,
) -> None:
    """Add print/export identity without disturbing row-1 data workflows."""
    branding = _resolve_branding(branding)
    title = _sheet_branding_title(report, sheet_label, branding)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(ws.max_column)}{max(header_row, ws.max_row)}"
    )
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.sheet_properties.tabColor = _xlsx_hex(tab_color)

    ws.oddHeader.left.text = title
    ws.oddHeader.left.font = "Aptos,Bold"
    ws.oddHeader.left.size = 9
    ws.oddHeader.left.color = _xlsx_hex(BRAND_INK)
    ws.oddHeader.right.text = "CONFIDENTIAL DRAFT - counsel review required"
    ws.oddHeader.right.font = "Aptos"
    ws.oddHeader.right.size = 8
    ws.oddHeader.right.color = _xlsx_hex(BRAND_SECONDARY_TEXT)
    if branding.suppresses_praviar_branding:
        ws.oddFooter.left.text = "AI-assisted FTO screening; not a legal opinion"
    else:
        ws.oddFooter.left.text = (
            "Generated by Praviar; AI-assisted FTO screening; not a legal opinion"
        )
    ws.oddFooter.left.font = "Aptos"
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = _xlsx_hex(BRAND_SECONDARY_TEXT)
    ws.oddFooter.right.text = "Page &[Page] of &[Pages]"
    ws.oddFooter.right.font = "Aptos"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = _xlsx_hex(BRAND_SECONDARY_TEXT)


def _set_header_row(ws, headers: list[str], *, row: int = 1) -> None:
    """Write a styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT_PAPER
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP


def _write_workbook_summary(
    wb: Workbook,
    report: FTOReport,
    options: ExportRenderOptions | None = None,
    branding: BrandingConfig | None = None,
) -> Worksheet:
    """Write visible Praviar branding as the workbook's first impression."""
    from praviar_pipeline.rendering.export_options import default_export_options

    options = options or default_export_options()
    branding = _resolve_branding(branding)
    ws = wb.create_sheet(title="Workbook Summary")
    wb.move_sheet(ws, offset=-(len(wb.sheetnames) - 1))
    wb.active = 0

    logo_path = resolve_branding_logo_path(
        branding,
        renderer_name="XLSX",
        supported_extensions=SUPPORTED_OFFICE_BRANDING_LOGO_EXTENSIONS,
    )
    if logo_path is not None:
        ws.add_image(OpenpyxlImage(str(logo_path)), "A1")
    elif not branding.suppresses_praviar_branding:
        ws.add_image(
            OpenpyxlImage(render_praviar_mark_png_stream(variant="on_light", size_px=256)),
            "A1",
        )
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 10
    ws.column_dimensions["A"].width = 14
    ws.merge_cells("B1:H1")
    ws.merge_cells("B2:H2")
    ws.merge_cells("B4:H4")
    ws.merge_cells("B5:H5")
    ws.merge_cells("B6:H6")
    ws.merge_cells("B7:H7")

    title_cell = ws.cell(row=1, column=2, value=_workbook_title(branding))
    title_cell.font = Font(bold=True, size=16, color=_xlsx_hex(BRAND_INK))
    title_cell.alignment = Alignment(vertical="center")

    subtitle_cell = ws.cell(
        row=2,
        column=2,
        value=f"{report.compound.name} | Report ID: {report.report_id}",
    )
    subtitle_cell.font = Font(size=10, color=_xlsx_hex(BRAND_SECONDARY_TEXT))
    subtitle_cell.alignment = Alignment(vertical="center")

    accent_cell = ws.cell(row=3, column=2)
    accent_cell.fill = PatternFill(
        start_color=_xlsx_hex(BRAND_COPPER),
        end_color=_xlsx_hex(BRAND_COPPER),
        fill_type="solid",
    )

    summary_cell = ws.cell(
        row=4,
        column=2,
        value=(
            f"Clearance decision: {governed_decision_label(report)} | "
            f"Patents analyzed: {report.risk_summary.total_patents_analyzed} | "
            f"Blocking patents: {governed_blocking_count(report)}"
        ),
    )
    summary_cell.font = Font(bold=True, size=11, color=_xlsx_hex(BRAND_INK))

    note_cell = ws.cell(
        row=5,
        column=2,
        value=(
            f"Audience: {options.audience_label} | Sections: {', '.join(options.section_labels)}"
        ),
    )
    note_cell.font = Font(size=10, color=_xlsx_hex(BRAND_SECONDARY_TEXT))

    workflow_cell = ws.cell(
        row=6,
        column=2,
        value=(
            "Structured sheets preserve source caveats for counsel review and downstream workflows."
        ),
    )
    workflow_cell.font = Font(size=10, color=_xlsx_hex(BRAND_SECONDARY_TEXT))

    schema_cell = ws.cell(
        row=7,
        column=2,
        value=f"Audience schema: {AUDIENCE_PROJECTION_SCHEMA_VERSION}",
    )
    schema_cell.font = Font(size=9, color=_xlsx_hex(BRAND_SECONDARY_TEXT))

    return ws


def _write_legal_notice(
    wb: Workbook,
    report: FTOReport,
    options: ExportRenderOptions,
    branding: BrandingConfig | None = None,
) -> Worksheet:
    """Write immutable legal marking and disclaimer as machine-visible cells."""
    ws = wb.create_sheet(title="Legal Notice", index=1)
    _set_header_row(ws, ["Field", "Value"])
    rows = (
        ("Legal Marking", "CONFIDENTIAL DRAFT"),
        ("Audience", options.audience_label),
        ("Audience Schema Version", AUDIENCE_PROJECTION_SCHEMA_VERSION),
        ("Disclaimer", REPORT_DISCLAIMER),
    )
    for row_number, (field, value) in enumerate(rows, 2):
        ws.cell(row=row_number, column=1, value=field).font = Font(
            bold=True, color=_xlsx_hex(BRAND_INK)
        )
        value_cell = ws.cell(row=row_number, column=2, value=value)
        value_cell.alignment = _WRAP
    _apply_working_sheet_branding(
        ws,
        report,
        sheet_label="Legal Notice",
        tab_color=BRAND_COPPER,
        branding=branding,
    )
    _auto_width(ws)
    return ws


def _write_source_audit(
    wb: Workbook,
    report: FTOReport,
    branding: BrandingConfig | None = None,
) -> Worksheet:
    """Write visible source-health and evidence-scope caveats."""
    ws = wb.create_sheet(title="Source Audit", index=2)
    summary = summarize_evidence_scope(report)

    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="Evidence Scope & Source Audit")
    title_cell.font = Font(bold=True, size=15, color=_xlsx_hex(BRAND_INK))
    title_cell.alignment = Alignment(vertical="center")

    ws.merge_cells("A2:D2")
    posture_cell = ws.cell(
        row=2,
        column=1,
        value=f"{summary.headline}. {summary.posture}.",
    )
    posture_cell.font = Font(bold=True, size=10, color=_xlsx_hex(BRAND_TEAL))
    posture_cell.alignment = _WRAP

    summary_rows = [
        ("Confidence impact", summary.confidence_impact),
        ("Counsel review note", summary.review_note),
        (
            "Successful sources",
            ", ".join(summary.successful_sources) or "No telemetry recorded",
        ),
        ("Unavailable sources", ", ".join(summary.unavailable_sources) or "None recorded"),
        ("Skipped sources", ", ".join(summary.skipped_sources) or "None recorded"),
    ]
    for row, (label, value) in enumerate(summary_rows, 4):
        ws.cell(row=row, column=1, value=label).font = Font(
            bold=True,
            color=_xlsx_hex(BRAND_INK),
        )
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.alignment = _WRAP
        value_cell.font = Font(color=_xlsx_hex(BRAND_SECONDARY_TEXT))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    header_row = 11
    _set_header_row(ws, ["Source", "Status", "Patents", "Detail"], row=header_row)
    if report.source_health.entries:
        for row, entry in enumerate(report.source_health.entries, header_row + 1):
            ws.cell(row=row, column=1, value=entry.source)
            status_cell = ws.cell(row=row, column=2, value=source_status_label(entry))
            if entry.status.value in {"failed", "not_configured"}:
                status_cell.fill = PatternFill(
                    start_color=_xlsx_hex(RISK_BG[RiskLevel.MEDIUM]),
                    end_color=_xlsx_hex(RISK_BG[RiskLevel.MEDIUM]),
                    fill_type="solid",
                )
            elif entry.status.value == "ok":
                status_cell.fill = PatternFill(
                    start_color=_xlsx_hex(RISK_BG[RiskLevel.CLEAR]),
                    end_color=_xlsx_hex(RISK_BG[RiskLevel.CLEAR]),
                    fill_type="solid",
                )
            ws.cell(row=row, column=3, value=entry.patent_count)
            detail_cell = ws.cell(row=row, column=4, value=source_status_detail(entry))
            detail_cell.alignment = _WRAP
    else:
        ws.cell(
            row=header_row + 1,
            column=1,
            value="Source-health telemetry not recorded",
        )
        ws.merge_cells(
            start_row=header_row + 1,
            start_column=1,
            end_row=header_row + 1,
            end_column=4,
        )

    _apply_working_sheet_branding(
        ws,
        report,
        header_row=header_row,
        sheet_label="Source Audit",
        tab_color=BRAND_TEAL,
        branding=branding,
    )
    _auto_width(ws)
    return ws


def _auto_width(ws, min_width: int = 12, max_width: int = 60) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                longest_line = max(len(line) for line in lines)
                max_len = max(max_len, min(longest_line + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _write_risk_matrix(
    wb: Workbook,
    report: FTOReport,
    branding: BrandingConfig | None = None,
    *,
    options: ExportRenderOptions | None = None,
) -> None:
    """Write the risk matrix sheet."""
    from praviar_pipeline.rendering.export_options import default_export_options

    options = options or default_export_options()
    ws: Worksheet = wb.active  # Always Worksheet for new Workbook
    ws.title = "Risk Matrix"

    headers = [
        "Patent ID",
        "Title",
        "Assignee",
        "Matter Clearance Decision",
        "Governed Patent Posture",
        "Claim-Coverage Screen",
        "Expiry Date",
    ]
    scientist_projection = options.audience == "scientist"
    if not scientist_projection:
        headers.extend(
            [
                "Claims Analyzed",
                "Orange Book",
                "Governed Basis",
                "Upstream Screen Summary",
            ]
        )
    header_row = 1
    _set_header_row(ws, headers, row=header_row)

    sorted_analyses = sorted(
        report.patent_analyses,
        key=lambda analysis: (
            {
                "BLOCKING": 0,
                "UNRESOLVED": 1,
                "NON-BLOCKING": 2,
                "SUPPORTING ONLY": 3,
            }.get(governed_patent_posture(report, analysis.patent_id), 4),
            {"high": 0, "medium": 1, "low": 2, "clear": 3}.get(
                analysis.risk_level.value,
                4,
            ),
            analysis.patent_id,
        ),
    )

    settings = get_settings()
    for row, a in enumerate(sorted_analyses, header_row + 1):
        ws.cell(row=row, column=1, value=a.patent_id)
        ws.cell(row=row, column=2, value=a.title[: settings.render_title_max_chars])
        ws.cell(row=row, column=3, value=a.assignee)
        ws.cell(row=row, column=4, value=governed_decision_label(report))
        posture = governed_patent_posture(report, a.patent_id)
        posture_cell = ws.cell(row=row, column=5, value=posture)
        posture_fill_key = {
            "BLOCKING": RiskLevel.HIGH.value,
            "UNRESOLVED": RiskLevel.MEDIUM.value,
            "NON-BLOCKING": RiskLevel.CLEAR.value,
            "SUPPORTING ONLY": RiskLevel.LOW.value,
        }.get(posture, RiskLevel.MEDIUM.value)
        posture_cell.fill = _RISK_FILLS.get(posture_fill_key, PatternFill())
        ws.cell(row=row, column=6, value=a.risk_level.value.upper())
        ws.cell(
            row=row,
            column=7,
            value=a.expiry_date.isoformat() if a.expiry_date else "Unknown",
        )
        claims = ", ".join(str(c.claim_number) for c in a.claims_analyzed)
        if not scientist_projection:
            ws.cell(row=row, column=8, value=claims)
            # Orange Book status is legal/regulatory detail reserved for counsel.
            ob_status = ""
            if a.orange_book_info and a.orange_book_info.is_listed:
                ob_status = (
                    "LISTED — DELIST REQUESTED" if a.orange_book_info.delist_requested else "LISTED"
                )
                if a.orange_book_info.nda_numbers:
                    ob_status += f" ({', '.join(a.orange_book_info.nda_numbers)})"
            ws.cell(row=row, column=9, value=ob_status)
            basis_cell = ws.cell(
                row=row,
                column=10,
                value=governed_patent_basis(report, a.patent_id),
            )
            basis_cell.alignment = _WRAP
            summary_cell = ws.cell(
                row=row,
                column=11,
                value=a.risk_summary[: settings.render_summary_max_chars],
            )
            summary_cell.alignment = _WRAP

    ws.freeze_panes = "A2"
    _apply_working_sheet_branding(
        ws,
        report,
        sheet_label="Risk Matrix",
        tab_color=BRAND_COPPER,
        branding=branding,
    )
    _auto_width(ws)


def _write_claim_charts(
    wb: Workbook,
    report: FTOReport,
    branding: BrandingConfig | None = None,
) -> None:
    """Write per-patent claim chart sheets."""
    used_titles = set(wb.sheetnames)
    for ia in report.invalidity_assessments:
        if not ia.claim_charts:
            continue

        sheet_name = _safe_sheet_title("CC", ia.patent_id, used_titles)
        ws = wb.create_sheet(title=sheet_name)

        headers = [
            "Patent ID",
            "Claim #",
            "Element #",
            "Element Text",
            "Prior Art Ref",
            "Prior Art Disclosure",
            "Citation Location",
            "Disclosed",
            "Notes",
        ]
        _set_header_row(ws, headers)

        row = 2
        for chart in ia.claim_charts:
            for entry in chart.entries:
                ws.cell(row=row, column=1, value=chart.patent_id)
                ws.cell(row=row, column=2, value=chart.claim_number)
                ws.cell(row=row, column=3, value=entry.element_number)

                elem_cell = ws.cell(row=row, column=4, value=entry.element_text)
                elem_cell.alignment = _WRAP

                ws.cell(row=row, column=5, value=entry.prior_art_reference_id)

                disc_cell = ws.cell(
                    row=row,
                    column=6,
                    value=entry.prior_art_disclosure,
                )
                disc_cell.alignment = _WRAP

                ws.cell(row=row, column=7, value=entry.citation_location)

                disclosed_cell = ws.cell(
                    row=row,
                    column=8,
                    value=entry.disclosed.upper(),
                )
                disclosed_cell.fill = _DISCLOSED_FILLS.get(
                    entry.disclosed,
                    PatternFill(),
                )

                notes_cell = ws.cell(row=row, column=9, value=entry.notes)
                notes_cell.alignment = _WRAP

                row += 1

            # Add chart summary as a merged row
            if chart.chart_summary:
                ws.cell(row=row, column=1, value="Summary:")
                summary_cell = ws.cell(row=row, column=2, value=chart.chart_summary)
                summary_cell.alignment = _WRAP
                summary_cell.font = Font(italic=True)
                row += 1

            # Blank row between charts
            row += 1

        _apply_working_sheet_branding(
            ws,
            report,
            sheet_label=f"Claim Chart {ia.patent_id}",
            tab_color=BRAND_MINT,
            branding=branding,
        )
        _auto_width(ws)


def _write_prior_art(
    wb: Workbook,
    report: FTOReport,
    branding: BrandingConfig | None = None,
) -> None:
    """Write a consolidated prior art references sheet."""
    # Collect all unique references
    seen_refs: set[str] = set()
    all_refs = []
    for ia in report.invalidity_assessments:
        for ref in ia.prior_art:
            if ref.reference_id not in seen_refs:
                seen_refs.add(ref.reference_id)
                all_refs.append((ia.patent_id, ref))

    if not all_refs:
        return

    settings = get_settings()
    ws = wb.create_sheet(title="Prior Art References")
    headers = [
        "Blocking Patent",
        "Reference ID",
        "Type",
        "Title",
        "Publication Date",
        "Anticipation Score",
        "Obviousness Score",
        "Authors",
        "Journal/Source",
        "DOI",
    ]
    _set_header_row(ws, headers)

    for row, (blocking_pid, ref) in enumerate(all_refs, 2):
        ws.cell(row=row, column=1, value=blocking_pid)
        ws.cell(row=row, column=2, value=ref.reference_id)
        ws.cell(row=row, column=3, value=ref.reference_type)
        ws.cell(row=row, column=4, value=ref.title[: settings.render_title_max_chars])
        ws.cell(
            row=row,
            column=5,
            value=ref.publication_date.isoformat() if ref.publication_date else "",
        )
        ws.cell(row=row, column=6, value=round(ref.anticipation_score, 2))
        ws.cell(row=row, column=7, value=round(ref.obviousness_score, 2))
        ws.cell(row=row, column=8, value=", ".join(ref.authors[:5]))
        ws.cell(row=row, column=9, value=ref.journal or ref.source_database)
        ws.cell(row=row, column=10, value=ref.doi)

    _apply_working_sheet_branding(
        ws,
        report,
        sheet_label="Prior Art References",
        tab_color=BRAND_TEAL,
        branding=branding,
    )
    _auto_width(ws)


def render_xlsx(
    report: FTOReport,
    options: ExportRenderOptions | None = None,
    branding: BrandingConfig | None = None,
) -> bytes:
    """Render an FTO report as an XLSX workbook.

    Returns the workbook as bytes (write to file or send as HTTP response).
    """
    from praviar_pipeline.rendering.export_options import default_export_options

    options = options or default_export_options()
    report = type(report).model_validate(
        neutralize_spreadsheet_structure(report.model_dump(mode="python"))
    )
    branding = _resolve_branding(branding)
    wb = Workbook()

    if options.allows(AudienceField.PATENT_LANDSCAPE) and options.includes("patent_analysis"):
        _write_risk_matrix(wb, report, branding, options=options)
    else:
        wb.remove(wb.active)

    _write_workbook_summary(wb, report, options, branding)
    _write_legal_notice(wb, report, options, branding)
    _write_source_audit(wb, report, branding)
    if options.allows(AudienceField.CLAIM_CHARTS) and options.includes("claim_charts"):
        _write_claim_charts(wb, report, branding)
    if options.allows(AudienceField.INVALIDITY_DETAIL) and options.includes(
        "invalidity_assessment"
    ):
        _write_prior_art(wb, report, branding)

    wb.properties.keywords = (
        f"CONFIDENTIAL DRAFT; {AUDIENCE_PROJECTION_SCHEMA_VERSION}; audience={options.audience}"
    )
    wb.properties.description = REPORT_DISCLAIMER

    neutralize_workbook_strings(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
