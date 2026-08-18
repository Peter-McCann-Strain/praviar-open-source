"""Tests for XLSX rendering — risk matrix, claim charts, prior art sheets."""

from __future__ import annotations

import base64
import io
from datetime import date

import pytest
from openpyxl import load_workbook

from praviar_pipeline.models.analysis import (
    ClaimAnalysis,
    ClaimElement,
    ElementStatus,
    PatentAnalysis,
    RiskLevel,
)
from praviar_pipeline.models.compound import ResolvedCompound
from praviar_pipeline.models.invalidity import (
    ClaimChart,
    ClaimChartEntry,
    InvalidityAssessment,
    PriorArtReference,
    PTABResult,
)
from praviar_pipeline.models.patent_term_models import OrangeBookInfo
from praviar_pipeline.models.report import (
    FTOReport,
    RiskSummary,
    SourceHealth,
    SourceHealthEntry,
    SourceStatus,
)
from praviar_pipeline.rendering.branding import BrandingConfig
from praviar_pipeline.rendering.export_options import ExportRenderOptions
from praviar_pipeline.rendering.xlsx import (
    _apply_working_sheet_branding,
    _auto_width,
    _safe_sheet_title,
    _set_header_row,
    _write_claim_charts,
    _write_prior_art,
    _write_risk_matrix,
    _write_source_audit,
    _write_workbook_summary,
    render_xlsx,
)

_PNG_1X1 = (
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
)

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _minimal_compound() -> ResolvedCompound:
    return ResolvedCompound(
        name="aspirin",
        canonical_smiles="CC(=O)Oc1ccccc1C(=O)O",
        inchi="InChI=1S/C9H8O4",
        inchi_key="BSYNRYMUTXBXSQ-UHFFFAOYSA-N",
        pubchem_cid=2244,
        molecular_formula="C9H8O4",
        molecular_weight=180.16,
        original_input="aspirin",
        input_type="name",
    )


def _write_png_logo(path) -> None:
    path.write_bytes(base64.b64decode(_PNG_1X1))


def _workbook_visible_text(wb) -> str:
    parts: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    parts.append(str(cell.value))
        for item in (
            ws.oddHeader.left.text,
            ws.oddHeader.center.text,
            ws.oddHeader.right.text,
            ws.oddFooter.left.text,
            ws.oddFooter.center.text,
            ws.oddFooter.right.text,
        ):
            if item:
                parts.append(str(item))
    return "\n".join(parts)


def _minimal_analysis(
    patent_id: str = "US1234567B2",
    risk_level: RiskLevel = RiskLevel.HIGH,
) -> PatentAnalysis:
    return PatentAnalysis(
        patent_id=patent_id,
        title="A test patent on aspirin synthesis",
        assignee="Pharma Corp",
        expiry_date=date(2030, 6, 1),
        claims_analyzed=[
            ClaimAnalysis(
                claim_number=1,
                claim_type="independent",
                elements=[
                    ClaimElement(
                        element_number=1,
                        element_text="A compound comprising acetylsalicylic acid",
                        status=ElementStatus.MET,
                        reasoning="Target is acetylsalicylic acid",
                        confidence=0.95,
                    ),
                ],
                overall_status=ElementStatus.MET,
                overall_confidence=0.95,
            )
        ],
        risk_level=risk_level,
        risk_summary="High risk — direct compound match",
    )


def _minimal_report(
    analyses: list[PatentAnalysis] | None = None,
    invalidity_assessments=None,
    source_health: SourceHealth | None = None,
) -> FTOReport:
    if analyses is None:
        analyses = [_minimal_analysis()]
    return FTOReport(
        report_id="xlsx-test-001",
        compound=_minimal_compound(),
        risk_summary=RiskSummary(
            overall_risk=RiskLevel.HIGH,
            blocking_patents_count=1,
            total_patents_analyzed=1,
            key_risks=["US1234567B2: high risk"],
            executive_summary="High FTO risk for aspirin.",
        ),
        patent_analyses=analyses,
        invalidity_assessments=invalidity_assessments or [],
        source_health=source_health or SourceHealth(),
    )


# ---------------------------------------------------------------------------
# Helper function unit tests
# ---------------------------------------------------------------------------


class TestSetHeaderRow:
    def test_header_cells_have_correct_values(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        headers = ["Patent ID", "Risk Level", "Expiry"]
        _set_header_row(ws, headers)
        assert ws.cell(row=1, column=1).value == "Patent ID"
        assert ws.cell(row=1, column=2).value == "Risk Level"
        assert ws.cell(row=1, column=3).value == "Expiry"

    def test_header_cells_have_brand_paper_font(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        _set_header_row(ws, ["A", "B"])
        cell = ws.cell(row=1, column=1)
        assert cell.font.color.rgb.upper().endswith("F6F4EF")

    def test_header_cells_have_dark_fill(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        _set_header_row(ws, ["X"])
        cell = ws.cell(row=1, column=1)
        assert cell.fill.start_color.rgb.upper().endswith("0B1F24")


class TestAutoWidth:
    def test_auto_width_sets_column_dimensions(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Short")
        ws.cell(row=2, column=1, value="A much longer text value here")
        _auto_width(ws)
        from openpyxl.utils import get_column_letter

        width = ws.column_dimensions[get_column_letter(1)].width
        assert width >= 12

    def test_auto_width_respects_max_width(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        # Insert a very long string
        ws.cell(row=1, column=1, value="x" * 200)
        _auto_width(ws, max_width=60)
        from openpyxl.utils import get_column_letter

        width = ws.column_dimensions[get_column_letter(1)].width
        assert width <= 60


class TestSafeSheetTitle:
    def test_replaces_excel_invalid_characters(self):
        used_titles: set[str] = set()

        title = _safe_sheet_title("CC", "WO/2026/123456?A1*[trial]", used_titles)

        assert title == "CC_WO_2026_123456_A1__trial_"
        assert len(title) <= 31
        assert used_titles == {title}

    def test_preserves_collision_resistance_after_truncation(self):
        used_titles: set[str] = set()
        first = _safe_sheet_title(
            "CC",
            "US123456789012345678901234567890A",
            used_titles,
        )
        second = _safe_sheet_title(
            "CC",
            "US123456789012345678901234567890B",
            used_titles,
        )

        assert first != second
        assert len(first) <= 31
        assert len(second) <= 31
        assert first.startswith("CC_US1234567890123456789")
        assert second.startswith("CC_US1234567890123456789")


# ---------------------------------------------------------------------------
# Risk matrix sheet tests
# ---------------------------------------------------------------------------


class TestWriteRiskMatrix:
    def test_risk_matrix_title(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report()
        _write_risk_matrix(wb, report)
        assert wb.active.title == "Risk Matrix"

    def test_risk_matrix_headers(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report()
        _write_risk_matrix(wb, report)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Patent ID"
        assert ws.cell(row=1, column=4).value == "Matter Clearance Decision"
        assert ws.cell(row=1, column=5).value == "Governed Patent Posture"
        assert ws.cell(row=1, column=6).value == "Claim-Coverage Screen"

    def test_risk_matrix_remains_machine_readable(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report()
        _write_risk_matrix(wb, report)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Patent ID"
        assert ws.cell(row=2, column=1).value == "US1234567B2"
        assert len(ws._images) == 0
        assert ws.freeze_panes == "A2"
        assert ws.print_title_rows == "$1:$1"
        assert ws.auto_filter.ref == ws.dimensions

    def test_risk_matrix_has_branded_print_header_and_footer(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report()
        _write_risk_matrix(wb, report)
        ws = wb.active

        assert ws.oddHeader.left.text == ("Praviar FTO Workbook | aspirin | Risk Matrix")
        assert ws.oddHeader.right.text == "CONFIDENTIAL DRAFT - counsel review required"
        assert ws.oddFooter.left.text == (
            "Generated by Praviar; AI-assisted FTO screening; not a legal opinion"
        )
        assert ws.oddFooter.right.text == "Page &[Page] of &[Pages]"
        assert ws.sheet_properties.tabColor.rgb.upper().endswith("B87333")

    def test_risk_matrix_data_row(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report()
        _write_risk_matrix(wb, report)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "US1234567B2"
        assert ws.cell(row=2, column=4).value == "UNCLEAR"
        assert ws.cell(row=2, column=5).value == "UNRESOLVED"
        assert ws.cell(row=2, column=6).value == "HIGH"

    def test_risk_matrix_sorts_by_risk(self):
        """HIGH risk patents appear before MEDIUM before LOW."""
        from openpyxl import Workbook

        analyses = [
            _minimal_analysis("USLOW", RiskLevel.LOW),
            _minimal_analysis("USHIGH", RiskLevel.HIGH),
            _minimal_analysis("USMED", RiskLevel.MEDIUM),
        ]
        report = _minimal_report(analyses=analyses)
        wb = Workbook()
        _write_risk_matrix(wb, report)
        ws = wb.active
        # Row 2 = first data row (highest risk) beneath the machine-readable header.
        assert ws.cell(row=2, column=1).value == "USHIGH"

    def test_orange_book_listed_included(self):
        """Patents listed in the Orange Book show status in the governed matrix."""
        from openpyxl import Workbook

        analysis = _minimal_analysis()
        analysis.orange_book_info = OrangeBookInfo(
            is_listed=True,
            nda_numbers=["NDA020563"],
            delist_requested=False,
        )
        report = _minimal_report(analyses=[analysis])
        wb = Workbook()
        _write_risk_matrix(wb, report)
        ws = wb.active
        ob_cell = ws.cell(row=2, column=9).value
        assert ob_cell is not None
        assert "LISTED" in ob_cell

    def test_unknown_expiry_renders_as_unknown(self):
        from openpyxl import Workbook

        analysis = _minimal_analysis()
        analysis.expiry_date = None
        report = _minimal_report(analyses=[analysis])
        wb = Workbook()
        _write_risk_matrix(wb, report)
        ws = wb.active
        assert ws.cell(row=2, column=7).value == "Unknown"


class TestWriteWorkbookSummary:
    def test_workbook_summary_contains_praviar_lockup(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report()
        _write_risk_matrix(wb, report)
        summary_ws = _write_workbook_summary(wb, report)

        assert summary_ws.title == "Workbook Summary"
        assert summary_ws.cell(row=1, column=2).value == "Praviar FTO Workbook"
        assert summary_ws.cell(row=2, column=2).value == ("aspirin | Report ID: xlsx-test-001")
        assert len(summary_ws._images) >= 1

    def test_workbook_summary_accepts_white_label_branding(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report()
        branding = BrandingConfig(hide_praviar_pipeline_branding=True)
        _write_risk_matrix(wb, report, branding)
        summary_ws = _write_workbook_summary(wb, report, branding=branding)

        assert summary_ws.cell(row=1, column=2).value == "FTO Analysis Workbook"
        assert len(summary_ws._images) == 0
        assert "Praviar" not in _workbook_visible_text(wb)

    def test_workbook_summary_accepts_custom_png_logo(self, tmp_path):
        from openpyxl import Workbook

        logo = tmp_path / "firm-logo.png"
        _write_png_logo(logo)
        branding = BrandingConfig(logo_path=str(logo), firm_name="Acme Counsel")

        wb = Workbook()
        report = _minimal_report()
        _write_risk_matrix(wb, report, branding)
        summary_ws = _write_workbook_summary(wb, report, branding=branding)

        assert summary_ws.cell(row=1, column=2).value == "Acme Counsel FTO Workbook"
        assert len(summary_ws._images) == 1
        assert "Praviar" not in _workbook_visible_text(wb)

    def test_workbook_summary_rejects_svg_logo_with_clear_error(self, tmp_path):
        from openpyxl import Workbook

        logo = tmp_path / "firm-logo.svg"
        logo.write_text("<svg xmlns='http://www.w3.org/2000/svg' />", encoding="utf-8")
        branding = BrandingConfig(logo_path=str(logo), firm_name="Acme Counsel")

        with pytest.raises(
            RuntimeError,
            match=r"XLSX does not support branding logo format '\.svg'",
        ):
            _write_workbook_summary(Workbook(), _minimal_report(), branding=branding)

    def test_workbook_summary_becomes_active_first_impression(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report()
        _write_risk_matrix(wb, report)
        _write_workbook_summary(wb, report)

        assert wb.active.title == "Workbook Summary"
        assert wb.worksheets[0].title == "Workbook Summary"
        assert wb["Risk Matrix"].cell(row=1, column=1).value == "Patent ID"
        assert wb["Risk Matrix"].cell(row=2, column=1).value == "US1234567B2"


class TestWriteSourceAudit:
    def test_source_audit_contains_front_matter_caveats(self):
        from openpyxl import Workbook

        health = SourceHealth(
            entries=[
                SourceHealthEntry(
                    source="pubchem",
                    status=SourceStatus.OK,
                    patent_count=42,
                ),
                SourceHealthEntry(
                    source="lens",
                    status=SourceStatus.NOT_CONFIGURED,
                    error_message="api_key=SUPERSECRET",
                ),
            ]
        )
        wb = Workbook()
        report = _minimal_report(source_health=health)
        _write_risk_matrix(wb, report)
        _write_workbook_summary(wb, report)
        audit_ws = _write_source_audit(wb, report)

        assert audit_ws.title == "Source Audit"
        assert audit_ws.cell(row=1, column=1).value == "Evidence Scope & Source Audit"
        assert "1 of 2 configured sources completed" in audit_ws.cell(row=2, column=1).value
        assert audit_ws.cell(row=4, column=1).value == "Confidence impact"
        assert audit_ws.cell(row=12, column=1).value == "pubchem"
        assert audit_ws.cell(row=13, column=2).value == "Not configured"
        assert "Provider was not configured" in audit_ws.cell(row=13, column=4).value
        assert "SUPERSECRET" not in audit_ws.cell(row=13, column=4).value

    def test_render_xlsx_includes_source_audit_second_sheet(self):
        health = SourceHealth(
            entries=[
                SourceHealthEntry(
                    source="pubchem",
                    status=SourceStatus.OK,
                    patent_count=42,
                ),
            ]
        )
        report = _minimal_report(source_health=health)

        wb = load_workbook(io.BytesIO(render_xlsx(report)))

        assert wb.sheetnames[:4] == [
            "Workbook Summary",
            "Legal Notice",
            "Source Audit",
            "Risk Matrix",
        ]
        assert wb["Source Audit"].oddHeader.left.text == (
            "Praviar FTO Workbook | aspirin | Source Audit"
        )


# ---------------------------------------------------------------------------
# Claim chart sheet tests
# ---------------------------------------------------------------------------


def _make_invalidity_with_chart(patent_id: str = "US1234567B2") -> InvalidityAssessment:
    chart = ClaimChart(
        patent_id=patent_id,
        claim_number=1,
        prior_art_reference_id="US5555555A",
        entries=[
            ClaimChartEntry(
                element_number=1,
                element_text="A compound comprising acetylsalicylic acid",
                prior_art_reference_id="US5555555A",
                prior_art_disclosure="Discloses acetylsalicylic acid compound",
                citation_location="col 3, lines 10-15",
                disclosed="yes",
                notes="Exact anticipation",
            ),
            ClaimChartEntry(
                element_number=2,
                element_text="wherein the purity is greater than 99%",
                prior_art_reference_id="US5555555A",
                prior_art_disclosure="Describes 98.5% pure compound",
                citation_location="col 4, line 20",
                disclosed="partial",
            ),
        ],
        chart_summary="Claim 1 substantially anticipated by US5555555A.",
    )
    return InvalidityAssessment(
        patent_id=patent_id,
        claim_numbers=[1],
        ptab=PTABResult(has_been_challenged=False),
        overall_invalidity_strength="moderate",
        reasoning="Prior art found",
        claim_charts=[chart],
    )


class TestWriteClaimCharts:
    def test_claim_chart_sheet_created(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report(invalidity_assessments=[_make_invalidity_with_chart()])
        _write_claim_charts(wb, report)
        sheet_names = [ws.title for ws in wb.worksheets]
        assert any("CC_" in name for name in sheet_names)

    def test_claim_chart_headers_correct(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report(invalidity_assessments=[_make_invalidity_with_chart()])
        _write_claim_charts(wb, report)
        # Find the claim chart sheet
        chart_ws = next(ws for ws in wb.worksheets if "CC_" in ws.title)
        assert chart_ws.cell(row=1, column=1).value == "Patent ID"
        assert chart_ws.cell(row=1, column=8).value == "Disclosed"
        assert chart_ws.freeze_panes == "A2"
        assert chart_ws.print_title_rows == "$1:$1"
        assert chart_ws.auto_filter.ref == chart_ws.dimensions
        assert chart_ws.oddHeader.left.text == (
            "Praviar FTO Workbook | aspirin | Claim Chart US1234567B2"
        )
        assert chart_ws.oddFooter.left.text == (
            "Generated by Praviar; AI-assisted FTO screening; not a legal opinion"
        )

    def test_claim_chart_data_rows(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report(invalidity_assessments=[_make_invalidity_with_chart()])
        _write_claim_charts(wb, report)
        chart_ws = next(ws for ws in wb.worksheets if "CC_" in ws.title)
        # Row 2 = first entry
        assert chart_ws.cell(row=2, column=1).value == "US1234567B2"
        assert chart_ws.cell(row=2, column=8).value == "YES"

    def test_claim_chart_partial_disclosed_present(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report(invalidity_assessments=[_make_invalidity_with_chart()])
        _write_claim_charts(wb, report)
        chart_ws = next(ws for ws in wb.worksheets if "CC_" in ws.title)
        # Row 3 = second entry (partial)
        assert chart_ws.cell(row=3, column=8).value == "PARTIAL"

    def test_chart_summary_row_added(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report(invalidity_assessments=[_make_invalidity_with_chart()])
        _write_claim_charts(wb, report)
        chart_ws = next(ws for ws in wb.worksheets if "CC_" in ws.title)
        # Summary row should appear after the data rows
        values = [chart_ws.cell(row=r, column=1).value for r in range(1, 10)]
        assert "Summary:" in values

    def test_no_claim_charts_no_sheet_created(self):
        """InvalidityAssessment with no claim charts should produce no extra sheet."""
        from openpyxl import Workbook

        ia = InvalidityAssessment(
            patent_id="US1234567B2",
            claim_numbers=[1],
            ptab=PTABResult(has_been_challenged=False),
            overall_invalidity_strength="weak",
            reasoning="No charts",
        )
        report = _minimal_report(invalidity_assessments=[ia])
        wb = Workbook()
        _write_claim_charts(wb, report)
        sheet_names = [ws.title for ws in wb.worksheets]
        assert not any("CC_" in name for name in sheet_names)

    def test_claim_chart_sheet_names_are_excel_safe_and_unique(self):
        from openpyxl import Workbook

        long_prefix = "US123456789012345678901234567890"
        assessments = [
            _make_invalidity_with_chart("WO/2026/123456?A1*[trial]"),
            _make_invalidity_with_chart(f"{long_prefix}A"),
            _make_invalidity_with_chart(f"{long_prefix}B"),
        ]
        wb = Workbook()
        report = _minimal_report(invalidity_assessments=assessments)

        _write_claim_charts(wb, report)

        chart_titles = [title for title in wb.sheetnames if title.startswith("CC_")]
        assert len(chart_titles) == 3
        assert len(set(chart_titles)) == 3
        assert all(len(title) <= 31 for title in chart_titles)
        assert all(not any(char in title for char in r":\/?*[]") for title in chart_titles)


# ---------------------------------------------------------------------------
# Prior art sheet tests
# ---------------------------------------------------------------------------


def _make_prior_art_ref(ref_id: str = "US5555555A") -> PriorArtReference:
    return PriorArtReference(
        reference_id=ref_id,
        title="Prior art on aspirin synthesis",
        publication_date=date(1995, 4, 10),
        reference_type="patent",
        anticipation_score=0.85,
        obviousness_score=0.6,
        authors=["Smith, J.", "Jones, K."],
        journal="",
        doi="",
    )


def _make_invalidity_with_prior_art(
    patent_id: str = "US1234567B2",
) -> InvalidityAssessment:
    return InvalidityAssessment(
        patent_id=patent_id,
        claim_numbers=[1],
        ptab=PTABResult(has_been_challenged=False),
        overall_invalidity_strength="moderate",
        reasoning="Strong prior art found",
        prior_art=[_make_prior_art_ref()],
    )


class TestWritePriorArt:
    def test_prior_art_sheet_created(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report(invalidity_assessments=[_make_invalidity_with_prior_art()])
        _write_prior_art(wb, report)
        sheet_names = [ws.title for ws in wb.worksheets]
        assert "Prior Art References" in sheet_names

    def test_prior_art_data_row(self):
        from openpyxl import Workbook

        wb = Workbook()
        report = _minimal_report(invalidity_assessments=[_make_invalidity_with_prior_art()])
        _write_prior_art(wb, report)
        ws = wb.worksheets[-1]
        assert ws.cell(row=2, column=2).value == "US5555555A"
        assert ws.cell(row=2, column=6).value == pytest.approx(0.85)
        assert ws.freeze_panes == "A2"
        assert ws.print_title_rows == "$1:$1"
        assert ws.oddHeader.left.text == ("Praviar FTO Workbook | aspirin | Prior Art References")

    def test_apply_working_sheet_branding_preserves_row_one_contract(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Patent ID")
        ws.cell(row=2, column=1, value="US1234567B2")
        report = _minimal_report()

        _apply_working_sheet_branding(
            ws,
            report,
            sheet_label="Counsel Worksheet",
        )

        assert ws.cell(row=1, column=1).value == "Patent ID"
        assert ws.cell(row=2, column=1).value == "US1234567B2"
        assert len(ws._images) == 0
        assert ws.freeze_panes == "A2"
        assert ws.oddHeader.left.text == ("Praviar FTO Workbook | aspirin | Counsel Worksheet")

    def test_prior_art_deduplication(self):
        """Same reference_id appearing in two invalidity assessments should produce one row."""
        from openpyxl import Workbook

        ia1 = _make_invalidity_with_prior_art("US1111111B2")
        ia2 = InvalidityAssessment(
            patent_id="US2222222B2",
            claim_numbers=[1],
            ptab=PTABResult(has_been_challenged=False),
            overall_invalidity_strength="weak",
            reasoning="Same prior art",
            prior_art=[_make_prior_art_ref("US5555555A")],  # same ref_id
        )
        report = _minimal_report(invalidity_assessments=[ia1, ia2])
        wb = Workbook()
        _write_prior_art(wb, report)
        ws = wb.worksheets[-1]
        # Only one data row
        refs = [
            ws.cell(row=r, column=2).value for r in range(2, 10) if ws.cell(row=r, column=2).value
        ]
        assert refs.count("US5555555A") == 1

    def test_no_prior_art_no_sheet(self):
        """When no invalidity assessments have prior art, no extra sheet is written."""
        from openpyxl import Workbook

        report = _minimal_report(invalidity_assessments=[])
        wb = Workbook()
        initial_count = len(wb.worksheets)
        _write_prior_art(wb, report)
        assert len(wb.worksheets) == initial_count


# ---------------------------------------------------------------------------
# End-to-end render_xlsx tests
# ---------------------------------------------------------------------------


class TestRenderXlsx:
    def test_returns_bytes(self):
        report = _minimal_report()
        result = render_xlsx(report)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_every_string_cell_neutralizes_hidden_formula_prefixes(self):
        report = _minimal_report()
        payload = ' \t=HYPERLINK("https://evil.example")'
        report.compound.name = payload
        report.patent_analyses[0].title = payload
        report.patent_analyses[0].assignee = "\r@SUM(1,2)"
        report.patent_analyses[0].risk_summary = "\x00+1+1"

        wb = load_workbook(io.BytesIO(render_xlsx(report)), data_only=False)

        visible = _workbook_visible_text(wb)
        assert "' \t=HYPERLINK" in visible
        assert "'\r@SUM" in visible
        assert "\N{REPLACEMENT CHARACTER}+1+1" in visible
        assert "\x00" not in visible
        for worksheet in wb.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    assert cell.data_type != "f"
                    if isinstance(cell.value, str):
                        probe = cell.value.lstrip(" \t\r\n\x00")
                        assert not probe.startswith(("=", "+", "-", "@")), cell.value

    def test_valid_xlsx_bytes(self):
        """Bytes should be parseable by openpyxl."""
        report = _minimal_report()
        data = render_xlsx(report)
        wb = load_workbook(io.BytesIO(data))
        assert len(wb.worksheets) >= 1

    def test_workbook_summary_uses_clearance_decision(self):
        report = _minimal_report()
        report.risk_summary.overall_risk = RiskLevel.CLEAR
        wb = load_workbook(io.BytesIO(render_xlsx(report)))

        summary = wb["Workbook Summary"].cell(row=4, column=2).value
        assert summary.startswith("Clearance decision: UNCLEAR")
        assert "Overall risk: CLEAR" not in summary

    def test_risk_matrix_sheet_present(self):
        report = _minimal_report()
        data = render_xlsx(report)
        wb = load_workbook(io.BytesIO(data))
        assert wb.active.title == "Workbook Summary"
        assert wb.worksheets[0].title == "Workbook Summary"
        assert "Risk Matrix" in wb.sheetnames

    def test_full_report_with_claim_charts_and_prior_art(self):
        ia = _make_invalidity_with_chart()
        ia.prior_art = [_make_prior_art_ref()]
        report = _minimal_report(invalidity_assessments=[ia])
        data = render_xlsx(report)
        wb = load_workbook(io.BytesIO(data))
        titles = [ws.title for ws in wb.worksheets]
        assert "Risk Matrix" in titles
        assert any("CC_" in t for t in titles)
        assert "Prior Art References" in titles
        assert wb["Risk Matrix"].oddHeader.left.text == (
            "Praviar FTO Workbook | aspirin | Risk Matrix"
        )
        assert wb["Risk Matrix"].cell(row=1, column=1).value == "Patent ID"
        assert len(wb["Risk Matrix"]._images) == 0

    def test_white_label_workbook_has_no_visible_praviar_branding(self):
        ia = _make_invalidity_with_chart()
        ia.prior_art = [_make_prior_art_ref()]
        report = _minimal_report(invalidity_assessments=[ia])
        branding = BrandingConfig(hide_praviar_pipeline_branding=True)

        wb = load_workbook(io.BytesIO(render_xlsx(report, branding=branding)))

        assert wb["Workbook Summary"].cell(row=1, column=2).value == ("FTO Analysis Workbook")
        assert "Praviar" not in _workbook_visible_text(wb)

    def test_missing_custom_logo_fails_closed(self, tmp_path):
        branding = BrandingConfig(
            logo_path=str(tmp_path / "missing-logo.png"),
            firm_name="Acme Counsel",
        )

        with pytest.raises(RuntimeError, match="Branding logo not found for XLSX"):
            render_xlsx(_minimal_report(), branding=branding)

    def test_scoped_workbook_omits_deselected_detail_sheets(self):
        ia = _make_invalidity_with_chart()
        ia.prior_art = [_make_prior_art_ref()]
        report = _minimal_report(invalidity_assessments=[ia])
        options = ExportRenderOptions.from_values(
            ["executive_summary"],
            audience="executive",
        )

        data = render_xlsx(report, options=options)
        wb = load_workbook(io.BytesIO(data))

        assert wb.sheetnames == ["Workbook Summary", "Legal Notice", "Source Audit"]
        assert "Audience: Executive Brief" in wb["Workbook Summary"]["B5"].value

    def test_metadata_only_scope_still_produces_valid_workbook(self):
        report = _minimal_report()
        options = ExportRenderOptions.from_values(
            ["pipeline_metadata"],
            audience="attorney",
        )

        data = render_xlsx(report, options=options)
        wb = load_workbook(io.BytesIO(data))

        assert wb.sheetnames == ["Workbook Summary", "Legal Notice", "Source Audit"]

    def test_empty_report_produces_valid_xlsx(self):
        """Report with no analyses should still produce a parseable workbook."""
        report = FTOReport(
            report_id="empty",
            compound=_minimal_compound(),
            risk_summary=RiskSummary(
                overall_risk=RiskLevel.CLEAR,
                blocking_patents_count=0,
                total_patents_analyzed=0,
                key_risks=[],
                executive_summary="Clear",
            ),
            patent_analyses=[],
        )
        data = render_xlsx(report)
        wb = load_workbook(io.BytesIO(data))
        assert wb.worksheets[0].title == "Workbook Summary"
        assert "Risk Matrix" in wb.sheetnames
