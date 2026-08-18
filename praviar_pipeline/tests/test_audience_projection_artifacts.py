"""Cross-format pairwise audience leakage tests for versioned projections."""

from __future__ import annotations

import io
from datetime import date

import pytest
from openpyxl import load_workbook

from praviar_pipeline.models.analysis import (
    ClaimAnalysis,
    ClaimElement,
    ElementStatus,
    PatentAnalysis,
    RiskLevel,
)
from praviar_pipeline.models.compound import ResolvedCompound
from praviar_pipeline.models.report import FTOReport, RiskSummary
from praviar_pipeline.rendering.audience_projection import (
    AUDIENCE_PROJECTION_SCHEMA_VERSION,
    AudienceField,
    audience_projection_policy,
)
from praviar_pipeline.rendering.csv import render_csv
from praviar_pipeline.rendering.docx_report import render_docx
from praviar_pipeline.rendering.export_options import ExportRenderOptions
from praviar_pipeline.rendering.pptx_report import render_pptx
from praviar_pipeline.rendering.xlsx import render_xlsx

PATENT_SENTINEL = "US-AUDIENCE-777B2"
ATTORNEY_DETAIL_SENTINEL = "ATTORNEY-ONLY-EVIDENCE-DETAIL-777"


@pytest.fixture
def audience_report() -> FTOReport:
    analysis = PatentAnalysis(
        patent_id=PATENT_SENTINEL,
        title="Audience projection patent",
        assignee="Projection Pharma",
        expiry_date=date(2035, 1, 1),
        claims_analyzed=[
            ClaimAnalysis(
                claim_number=1,
                claim_type="independent",
                elements=[
                    ClaimElement(
                        element_number=1,
                        element_text="attorney claim limitation",
                        status=ElementStatus.MET,
                        reasoning=ATTORNEY_DETAIL_SENTINEL,
                        confidence=0.9,
                    )
                ],
                overall_status=ElementStatus.MET,
                overall_confidence=0.9,
            )
        ],
        risk_level=RiskLevel.HIGH,
        risk_summary=ATTORNEY_DETAIL_SENTINEL,
    )
    return FTOReport(
        report_id="audience-projection-report",
        compound=ResolvedCompound(
            name="projection compound",
            original_input="projection compound",
            input_type="name",
        ),
        risk_summary=RiskSummary(
            overall_risk=RiskLevel.HIGH,
            blocking_patents_count=1,
            total_patents_analyzed=1,
            executive_summary="Governed audience summary",
        ),
        patent_analyses=[analysis],
        patent_narratives={PATENT_SENTINEL: ATTORNEY_DETAIL_SENTINEL},
    )


def _options(audience: str) -> ExportRenderOptions:
    return ExportRenderOptions.from_values(audience=audience)


def _docx_text(payload: bytes) -> str:
    docx = pytest.importorskip("docx")
    document = docx.Document(io.BytesIO(payload))
    parts = [paragraph.text for paragraph in document.paragraphs]
    for table in document.tables:
        parts.extend(cell.text for row in table.rows for cell in row.cells)
    return "\n".join(parts)


def _pptx_text(payload: bytes) -> str:
    pptx = pytest.importorskip("pptx")
    presentation = pptx.Presentation(io.BytesIO(payload))
    parts: list[str] = []
    for slide in presentation.slides:
        for shape in slide.shapes:
            if hasattr(shape, "text"):
                parts.append(shape.text)
            if getattr(shape, "has_table", False):
                parts.extend(cell.text for row in shape.table.rows for cell in row.cells)
    return "\n".join(parts)


def _xlsx_text(payload: bytes) -> str:
    workbook = load_workbook(io.BytesIO(payload))
    return "\n".join(
        str(cell.value)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )


def _assert_pairwise_contract(text_by_audience: dict[str, str]) -> None:
    assert ATTORNEY_DETAIL_SENTINEL in text_by_audience["attorney"]
    assert PATENT_SENTINEL in text_by_audience["attorney"]
    for audience in ("executive", "investor"):
        assert ATTORNEY_DETAIL_SENTINEL not in text_by_audience[audience]
        assert PATENT_SENTINEL not in text_by_audience[audience]
    assert ATTORNEY_DETAIL_SENTINEL not in text_by_audience["scientist"]
    assert PATENT_SENTINEL in text_by_audience["scientist"]


def test_audience_projection_v1_policy_snapshot() -> None:
    executive = audience_projection_policy("executive")
    executive_snapshot = {
        "schema_version": executive.schema_version,
        "audience": executive.audience,
        "allowed_sections": sorted(executive.allowed_sections),
        "allowed_fields": sorted(field.value for field in executive.allowed_fields),
    }
    assert executive_snapshot == {
        "schema_version": "audience-projection-v1",
        "audience": "executive",
        "allowed_sections": ["executive_summary"],
        "allowed_fields": ["executive_summary", "recommendations", "source_audit"],
    }
    scientist = audience_projection_policy("scientist")
    assert scientist.schema_version == AUDIENCE_PROJECTION_SCHEMA_VERSION
    assert AudienceField.PATENT_LANDSCAPE in scientist.allowed_fields
    assert AudienceField.CLAIM_CHARTS not in scientist.allowed_fields
    for restricted in ("executive", "investor", "scientist"):
        policy = audience_projection_policy(restricted)
        assert AudienceField.PATENT_DETAIL not in policy.allowed_fields
        assert AudienceField.INVALIDITY_DETAIL not in policy.allowed_fields


def test_docx_pairwise_forbidden_content(audience_report: FTOReport) -> None:
    texts = {
        audience: _docx_text(render_docx(audience_report, options=_options(audience)))
        for audience in ("attorney", "executive", "investor", "scientist")
    }
    _assert_pairwise_contract(texts)


def test_pptx_pairwise_forbidden_content(audience_report: FTOReport) -> None:
    texts = {
        audience: _pptx_text(render_pptx(audience_report, options=_options(audience)))
        for audience in ("attorney", "executive", "investor", "scientist")
    }
    _assert_pairwise_contract(texts)


def test_csv_pairwise_forbidden_content_and_disclaimer(audience_report: FTOReport) -> None:
    texts = {
        audience: "\n".join(render_csv(audience_report, options=_options(audience)).values())
        for audience in ("attorney", "executive", "investor", "scientist")
    }
    _assert_pairwise_contract(texts)
    for text in texts.values():
        assert "CONFIDENTIAL DRAFT" in text
        assert AUDIENCE_PROJECTION_SCHEMA_VERSION in text
        assert "does NOT constitute legal advice" in text


def test_xlsx_pairwise_forbidden_content_and_disclaimer(audience_report: FTOReport) -> None:
    payloads = {
        audience: render_xlsx(audience_report, options=_options(audience))
        for audience in ("attorney", "executive", "investor", "scientist")
    }
    texts = {audience: _xlsx_text(payload) for audience, payload in payloads.items()}
    _assert_pairwise_contract(texts)
    for audience, payload in payloads.items():
        workbook = load_workbook(io.BytesIO(payload))
        assert "Legal Notice" in workbook.sheetnames
        assert workbook["Legal Notice"]["B2"].value == "CONFIDENTIAL DRAFT"
        assert workbook["Legal Notice"]["B4"].value == AUDIENCE_PROJECTION_SCHEMA_VERSION
        assert "does NOT constitute legal advice" in workbook["Legal Notice"]["B5"].value
        assert f"audience={audience}" in workbook.properties.keywords
