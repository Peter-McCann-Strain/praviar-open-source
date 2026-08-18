"""Convert binary data files (Excel, Parquet) to CSV for agent consumption.

One-time conversion — run once, agents read CSVs thereafter.
"""

from pathlib import Path

RESEARCH_ROOT = Path(__file__).resolve().parents[2]
VALIDATION_DIR = RESEARCH_ROOT / "validation"


def convert_xls(src: Path, dst: Path) -> None:
    """Convert .xls (OLE2) to CSV."""
    import xlrd

    wb = xlrd.open_workbook(str(src))
    for sheet_idx in range(wb.nsheets):
        sheet = wb.sheet_by_index(sheet_idx)
        suffix = f"_sheet{sheet_idx}" if wb.nsheets > 1 else ""
        out_path = dst.with_name(dst.stem + suffix + ".csv")
        with open(out_path, "w", encoding="utf-8") as f:
            for row_idx in range(sheet.nrows):
                values = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    val = str(cell.value).replace(",", ";").replace("\n", " ").strip()
                    values.append(val)
                f.write(",".join(values) + "\n")
        print(f"  → {out_path.name} ({sheet.nrows} rows × {sheet.ncols} cols)")


def convert_xlsx(src: Path, dst: Path) -> None:
    """Convert .xlsx to CSV."""
    import openpyxl

    wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        suffix = f"_{sheet_name}" if len(wb.sheetnames) > 1 else ""
        out_path = dst.with_name(dst.stem + suffix + ".csv")
        with open(out_path, "w", encoding="utf-8") as f:
            for row in sheet.iter_rows(values_only=True):
                values = [
                    str(v).replace(",", ";").replace("\n", " ").strip()
                    if v is not None
                    else ""
                    for v in row
                ]
                f.write(",".join(values) + "\n")
        print(f"  → {out_path.name}")
    wb.close()


def convert_parquet(src: Path, dst: Path, sample_rows: int = 2000) -> None:
    """Convert Parquet to CSV, sampling rows to keep size manageable."""
    import pyarrow.parquet as pq

    table = pq.read_table(str(src))
    total = table.num_rows
    print(f"  Total rows: {total}, columns: {table.num_columns}")

    # Write sampled version
    if total > sample_rows:

        indices = list(range(0, total, total // sample_rows))[:sample_rows]
        table_sample = table.take(indices)
    else:
        table_sample = table

    table_sample.to_pandas().to_csv(str(dst), index=False)
    print(f"  → {dst.name} ({table_sample.num_rows} rows)")

    # Also write column info
    info_path = dst.with_name(dst.stem + "_columns.txt")
    with open(info_path, "w") as f:
        for i, name in enumerate(table.column_names):
            col = table.column(i)
            f.write(f"{name}: {col.type} (nulls: {col.null_count})\n")
    print(f"  → {info_path.name}")


def main() -> None:
    conversions = [
        # WIPO Ritonavir Excel DB
        (
            VALIDATION_DIR
            / "ground-truth/wipo-landscape-reports/ritonavir-appendices/database-patent-families.xls",
            VALIDATION_DIR
            / "ground-truth/wipo-landscape-reports/ritonavir-appendices/database-patent-families.csv",
            "xls",
        ),
        # WIPO Atazanavir Excel DB
        (
            VALIDATION_DIR
            / "ground-truth/wipo-landscape-reports/atazanavir-appendices/database-patent-families.xls",
            VALIDATION_DIR
            / "ground-truth/wipo-landscape-reports/atazanavir-appendices/database-patent-families.csv",
            "xls",
        ),
        # PANORAMA parquet
        (
            VALIDATION_DIR / "external-datasets/panorama/panorama.parquet",
            VALIDATION_DIR / "external-datasets/panorama/panorama_sample.csv",
            "parquet",
        ),
        # USPTO PTE certificates
        (
            VALIDATION_DIR / "external-datasets/uspto-pte/pte-certificates.xls",
            VALIDATION_DIR / "external-datasets/uspto-pte/pte-certificates.csv",
            "xls",
        ),
        # USPTO PTE past 5 years
        (
            VALIDATION_DIR / "external-datasets/uspto-pte/pte-past-5-years.xlsx",
            VALIDATION_DIR / "external-datasets/uspto-pte/pte-past-5-years.csv",
            "xlsx",
        ),
    ]

    ok = 0
    fail = 0
    for src, dst, fmt in conversions:
        print(f"\nConverting: {src.name}")
        if not src.exists():
            print(f"  SKIP — file not found: {src}")
            fail += 1
            continue
        if dst.exists():
            print(f"  SKIP — already converted: {dst.name}")
            ok += 1
            continue
        try:
            if fmt == "xls":
                convert_xls(src, dst)
            elif fmt == "xlsx":
                convert_xlsx(src, dst)
            elif fmt == "parquet":
                convert_parquet(src, dst)
            ok += 1
        except Exception as e:
            print(f"  FAILED: {e}")
            fail += 1

    print(f"\nDone: {ok} converted, {fail} failed")


if __name__ == "__main__":
    main()
